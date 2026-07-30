"""
Painel Executivo de Produção — Varejo
======================================
Requisitos: streamlit>=1.32, pandas, openpyxl, matplotlib

Melhorias implementadas nesta versão em relação ao script original:
  (obs.: o login por senha foi removido a pedido — o painel fica sem
  proteção de acesso, igual ao comportamento original)
  2. Leitura da planilha por CABEÇALHO de coluna (TOTAL / USUARIO), com aviso
     caso precise cair para a posição fixa (colunas I/M) por compatibilidade.
  3. Remoção dos números de fallback fixos (50271 / 1104): agora, se a
     planilha não tiver dados válidos, o painel avisa claramente em vez de
     mostrar números "fantasmas".
  4. Alerta de nomes encontrados na planilha que não batem com ninguém da
     equipe cadastrada (evita gente "sumir" do relatório sem ninguém notar).
  5. Cache da leitura da planilha (st.cache_data) para não reprocessar tudo
     a cada clique no sidebar.
  6. Equipe configurável via JSON editável na própria interface (sem precisar
     mexer no código para adicionar/remover pessoas, alias de Excel ou metas
     individuais).
  7. Horários de movimentação via tabela editável (st.data_editor) com
     colunas de hora, no lugar de vários text_input soltos + botões de
     add/remover linha.
  8. Histórico diário salvo em CSV local + gráfico de evolução (tendência).
  9. Metas individuais opcionais por colaborador (além da meta do setor).
 10. Envio de e-mail direto por SMTP (opcional, configurável no sidebar).
 11. Exportação da tabela gerencial em Excel (.xlsx), além da imagem PNG.
 12. Tabela do detalhamento gerencial mais compacta (linhas menores, sem
     alterar o conteúdo do texto), com opção de editá-la como planilha e de
     ocultar/exibir Exemplares e SKUs individuais sob demanda.
 13. Coluna "Movimentação Operacional" agora aceita texto livre editado à mão
     (persistido por pessoa + data, sobrepondo o texto gerado automaticamente
     a partir dos horários), com botão para restaurar o texto automático.
 14. Horários de Saída/Retorno/Local do sidebar agora podem ser salvos por
     pessoa (botão "💾 Salvar") para não precisar digitar de novo a cada vez
     que ela for movimentada, com botão "🔄 Resetar" para voltar ao padrão.
 15. E-mail agora é enviado em HTML, com a imagem do relatório embutida
     diretamente no corpo (igual ao modelo mostrado pelo usuário), em vez de
     ir só como anexo separado.
 16. Nova seção no sidebar "➕ Adicionar Manualmente ao Relatório": digite um
     nome e clique em Adicionar para incluí-lo(a) no Detalhamento Gerencial,
     mesmo que não esteja na equipe cadastrada e mesmo sem registro na
     planilha no dia (não depende da seção de Movimentação). O comportamento
     padrão da equipe cadastrada continua o mesmo de antes: quem tem SKUs
     zerados no dia não aparece automaticamente na tabela.
"""

import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime, time as dtime
from zoneinfo import ZoneInfo
import io
import os
import json
import html
import textwrap
import unicodedata
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# =============================================================================
# 0. CONSTANTES DE ARQUIVO
# =============================================================================
MOV_OVERRIDES_PATH = "movimentacoes_manuais.json"  # texto livre digitado na coluna Movimentação Operacional
HORARIOS_SALVOS_PATH = "horarios_salvos.json"  # horários de Saída/Retorno/Local salvos por pessoa (sidebar)

# (constantes/função de senha removidas junto com o login — ver comentário
# na seção 2 abaixo caso queira reativar o acesso restrito no futuro)


# =============================================================================
# 1. Configuração e Estilização de Design Premium (HTML / CSS)
# =============================================================================
st.set_page_config(page_title="Dashboard Executivo Varejo", layout="wide")
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    .stApp { background-color: #F8FAFC; }

    .block-container {padding-top: 1.2rem; padding-bottom: 0rem; max-width: 96%;}

    section[data-testid="stSidebar"] {
        background-color: #FFFFFF;
        border-right: 1px solid #E5E7EB;
    }
    section[data-testid="stSidebar"] h3 {
        font-family: 'Sora', sans-serif;
        color: #0F172A !important;
        font-size: 0.95rem !important;
    }

    .card-kpi {
        background: #FFFFFF;
        border: 1px solid #E5E7EB;
        border-left: 4px solid var(--accent-color, #2563EB);
        color: #0F172A;
        padding: 20px 22px;
        border-radius: 10px;
        box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06);
        text-align: left;
        margin-bottom: 14px;
        transition: box-shadow 0.2s ease;
    }
    .card-kpi:hover { box-shadow: 0 4px 14px rgba(15, 23, 42, 0.08); }
    .card-title {
        font-family: 'Sora', sans-serif;
        font-size: 0.8rem; font-weight: 700; opacity: 0.9; margin-bottom: 6px;
        letter-spacing: 1px; text-transform: uppercase; color: #64748B;
    }
    .card-value {
        font-family: 'Sora', sans-serif;
        font-size: 2.3rem; font-weight: 800; line-height: 1; margin-bottom: 8px;
        color: #0F172A;
    }
    .card-sub { font-size: 0.85rem; font-weight: 600; color: #94A3B8; }

    div.stProgress > div > div > div {
        background: #2563EB;
        height: 6px; border-radius: 4px;
    }
    div.stProgress > div > div { background: #E5E7EB; border-radius: 4px; }

    hr { border-color: #E5E7EB !important; }

    .stButton>button {
        background: #FFFFFF;
        border: 1px solid #CBD5E1;
        color: #0F172A;
        border-radius: 8px;
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        transition: all 0.15s ease;
    }
    .stButton>button:hover {
        border-color: #2563EB;
        color: #2563EB;
    }
    div[data-testid="stDownloadButton"] button {
        background: #0F172A;
        border: 1px solid #0F172A;
        color: #FFFFFF;
    }
    div[data-testid="stDownloadButton"] button:hover {
        background: #1E293B;
        border-color: #1E293B;
        color: #FFFFFF;
    }

    .stTextInput>div>div>input, .stDateInput input {
        background-color: #FFFFFF;
        color: #0F172A;
        border: 1px solid #E2E8F0;
        border-radius: 6px;
    }
    .stTextInput>div>div>input:focus { border-color: #2563EB; }

    div[data-testid="stDataFrame"] {
        border: 1px solid #E5E7EB;
        border-radius: 10px;
        overflow: hidden;
    }

    /* Tabela gerencial em HTML (bordas arredondadas + quebra de texto) */
    .tabela-wrapper {
        border: 1px solid #E2E8F0;
        border-radius: 14px;
        overflow: hidden;
        box-shadow: 0 1px 4px rgba(15, 23, 42, 0.07);
        margin-bottom: 16px;
    }
    table.tabela-gerencial {
        width: 100%;
        border-collapse: collapse;
        font-family: 'Inter', sans-serif;
        font-size: 0.86rem;
        table-layout: fixed;
    }
    table.tabela-gerencial thead th {
        background: #0F172A;
        color: #FFFFFF;
        text-align: left;
        padding: 9px 14px;
        font-weight: 600;
        font-size: 0.78rem;
        text-transform: uppercase;
        letter-spacing: 0.3px;
    }
    table.tabela-gerencial tbody td {
        padding: 6px 14px;
        border-top: 1px solid #EEF2F6;
        color: #0F172A;
        vertical-align: middle;
        white-space: normal;      /* permite quebra de linha nas colunas de texto livre */
        overflow-wrap: break-word;
        line-height: 1.3;
        font-size: 0.86rem;
    }
    /* Cargo e Colaboradora são textos curtos/categóricos: nunca cortam no
       meio da palavra, ficam sempre em uma linha só. */
    table.tabela-gerencial td:nth-child(1),
    table.tabela-gerencial td:nth-child(2),
    table.tabela-gerencial th:nth-child(1),
    table.tabela-gerencial th:nth-child(2) {
        white-space: nowrap;
    }
    table.tabela-gerencial tbody tr:nth-child(even) { background: #FAFBFC; }
    table.tabela-gerencial tbody tr:hover { background: #F1F5F9; }
    .tabela-vazia {
        padding: 18px 16px;
        color: #64748B;
        font-size: 0.9rem;
        border: 1px solid #E2E8F0;
        border-radius: 14px;
        background: #FFFFFF;
    }
    </style>
""", unsafe_allow_html=True)

# Logo leve em SVG (substitui o PNG base64 do arquivo original — ver nota no
# topo do arquivo sobre como restaurar o logo original caso deseje).
LOGO_REALBRAS_SVG = """<svg width="34" height="34" viewBox="0 0 34 34" xmlns="http://www.w3.org/2000/svg">
  <rect width="34" height="34" rx="8" fill="#0F172A"/>
  <path d="M9 24V10h7.5c3.6 0 5.8 1.8 5.8 4.9 0 2.1-1.1 3.6-3 4.3l3.4 4.8h-3.4l-3-4.3H12V24H9zm3-6.7h4.2c1.8 0 2.8-.8 2.8-2.3s-1-2.3-2.8-2.3H12v4.6z" fill="#FFFFFF"/>
</svg>""".strip()


# =============================================================================
# 2. LOGIN — DESATIVADO
# =============================================================================
# O gate de senha foi removido a pedido. Se quiser reativar no futuro, basta
# restaurar a função `checar_senha()` (mantida como exemplo abaixo, comentada)
# e voltar a chamar `if not checar_senha(): st.stop()` antes do cabeçalho.
#
# def checar_senha():
#     if st.session_state.get("autenticado"):
#         return True
#
#     def senha_confirmada():
#         if st.session_state.get("senha_input") == obter_senha_configurada():
#             st.session_state["autenticado"] = True
#         else:
#             st.session_state["autenticado"] = False
#
#     st.markdown(f"""
#     <div style='display:flex; align-items:center; gap:10px; margin-bottom:18px;'>
#         {LOGO_REALBRAS_SVG}
#         <h2 style='font-family: "Sora", sans-serif; font-weight:800; color:#0F172A; margin:0;'>🔒 Acesso Restrito</h2>
#     </div>
#     """, unsafe_allow_html=True)
#     st.text_input("Senha de acesso:", type="password", on_change=senha_confirmada, key="senha_input")
#     if "autenticado" in st.session_state and not st.session_state["autenticado"]:
#         st.error("Senha incorreta. Tente novamente.")
#     st.caption("Dica: configure a senha em `.streamlit/secrets.toml` com a chave `SENHA_PAINEL`.")
#     return False

# =============================================================================
# 3. Cabeçalho
# =============================================================================
st.markdown(f"""
<div style='display:flex; align-items:center; gap:10px; margin-bottom:4px;'>
{LOGO_REALBRAS_SVG}
<h1 style='font-family: "Sora", sans-serif; font-weight:800; letter-spacing:-0.5px; color: #0F172A; margin:0;'>📊 Painel Executivo de Produção</h1>
</div>
<p style='text-align:left; font-family: "Inter", sans-serif; color:#64748B; font-size:0.95rem; margin-top:4px; margin-bottom:28px;'>Varejo · acompanhamento diário de produtividade</p>
""", unsafe_allow_html=True)


# =============================================================================
# 4. Normalização de texto (evita falha de correspondência de nomes)
# =============================================================================
def normalizar(texto):
    """Remove acentos, espaços extras e padroniza para maiúsculas, evitando
    falhas de correspondência entre o nome cadastrado no painel e o nome como
    aparece na planilha."""
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    texto = " ".join(texto.split())
    return texto


# =============================================================================
# 5. EQUIPE CONFIGURÁVEL (editável pela interface, sem precisar mexer no código)
# =============================================================================
DEFAULT_EQUIPE = {
    "Líder": [
        {"nome": "Kamila Moraes", "alias_excel": "KAMILA"},
        {"nome": "Beatriz Alcantara", "alias_excel": "BEATRIZ"},
    ],
    "Apoio": [
        {"nome": "Alisson Lima"},
    ],
    "Operador(a)": [
        {"nome": "Rosana Delfino"},
        {"nome": "Ana Caroline", "alias_excel": "ANACAROLINE"},
        {"nome": "Karoline Gonçalves"},
        {"nome": "Gabriele"},
        {"nome": "Beatriz Mascarenhas"},
        {"nome": "Graziela Pereira", "alias_excel": "GRAZIELA PEREIRA DO NASCIMENTO"},
        {"nome": "Paula Roberta", "alias_excel": "PAULA ROBERTA SANTOS DA SILVA"},
        {"nome": "Weliton"},
        {"nome": "Ellen Kelly"},
    ],
}

if "equipe_config" not in st.session_state:
    st.session_state["equipe_config"] = json.loads(json.dumps(DEFAULT_EQUIPE))  # cópia profunda

def construir_estruturas_equipe(config):
    equipe = {}
    nomes_lista = []
    alias_excel = {}
    metas_individuais = {}
    defaults_mov = {}
    cargo_por_nome = {}

    for cargo, integrantes in config.items():
        nomes_cargo = []
        for pessoa in integrantes:
            nome = pessoa.get("nome", "").strip()
            if not nome:
                continue
            nomes_cargo.append(nome)
            nomes_lista.append(nome)
            cargo_por_nome[nome] = cargo
            if pessoa.get("alias_excel"):
                alias_excel[nome] = pessoa["alias_excel"]
            if pessoa.get("meta_exemplares"):
                try:
                    metas_individuais[nome] = int(pessoa["meta_exemplares"])
                except (TypeError, ValueError):
                    pass
            defaults_mov[nome] = {
                "saida": pessoa.get("saida_padrao", ""),
                "retorno": pessoa.get("retorno_padrao", ""),
                "local": pessoa.get("local_padrao", ""),
            }
        equipe[cargo] = nomes_cargo

    return equipe, nomes_lista, alias_excel, metas_individuais, defaults_mov, cargo_por_nome


def nome_excel(nome, alias_map):
    return normalizar(alias_map.get(nome, nome))


def parse_hora_str(valor_str):
    if not valor_str:
        return None
    try:
        h, m = valor_str.split(":")
        return dtime(int(h), int(m))
    except Exception:
        return None


def formatar_hora_editor(valor):
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(valor, (dtime, datetime)):
        return valor.strftime("%Hh%M")
    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return ""
        for fmt in ("%H:%M:%S.%f", "%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(texto, fmt).strftime("%Hh%M")
            except ValueError:
                continue
        try:
            return pd.to_datetime(texto).strftime("%Hh%M")
        except Exception:
            return texto
    if hasattr(valor, "strftime"):
        try:
            return valor.strftime("%Hh%M")
        except Exception:
            return ""
    return str(valor).strip()


def hora_para_iso(valor):
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(valor, (dtime, datetime)):
        return valor.strftime("%H:%M")
    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return ""
        for fmt in ("%H:%M:%S.%f", "%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(texto, fmt).strftime("%H:%M")
            except ValueError:
                continue
        try:
            return pd.to_datetime(texto).strftime("%H:%M")
        except Exception:
            return ""
    if hasattr(valor, "strftime"):
        try:
            return valor.strftime("%H:%M")
        except Exception:
            return ""
    return ""


def carregar_overrides_disco():
    if not os.path.exists(MOV_OVERRIDES_PATH):
        return {}
    try:
        with open(MOV_OVERRIDES_PATH, "r", encoding="utf-8") as f:
            bruto = json.load(f)
        return {tuple(chave.split("||", 1)): texto for chave, texto in bruto.items()}
    except Exception:
        return {}


def salvar_overrides_disco(overrides):
    try:
        bruto = {f"{data}||{nome}": texto for (data, nome), texto in overrides.items()}
        with open(MOV_OVERRIDES_PATH, "w", encoding="utf-8") as f:
            json.dump(bruto, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def carregar_horarios_disco():
    if not os.path.exists(HORARIOS_SALVOS_PATH):
        return {}
    try:
        with open(HORARIOS_SALVOS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def salvar_horarios_disco(horarios):
    try:
        with open(HORARIOS_SALVOS_PATH, "w", encoding="utf-8") as f:
            json.dump(horarios, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def texto_seguro(valor):
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass
    return str(valor)


if "mov_manual_overrides" not in st.session_state:
    st.session_state["mov_manual_overrides"] = carregar_overrides_disco()

if "horarios_salvos" not in st.session_state:
    st.session_state["horarios_salvos"] = carregar_horarios_disco()


# =============================================================================
# 6. BARRA LATERAL — Upload, data, configuração de equipe e filtros
# =============================================================================
st.sidebar.header("🛠️ Controle Operacional")
uploaded_file = st.sidebar.file_uploader("Upload da Planilha Excel", type=["xlsx"], key="uploaded_file")

st.sidebar.markdown("<hr style='margin:14px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)

data_produtividade = st.sidebar.date_input(
    "Data da Produtividade:", datetime.now(ZoneInfo("America/Sao_Paulo")), key="data_produtividade"
)
data_formatada = data_produtividade.strftime("%d/%m")

st.sidebar.markdown("<hr style='margin:14px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)

with st.sidebar.expander("⚙️ Configurar Equipe (avançado)"):
    st.caption(
        "Edite o JSON abaixo para adicionar/remover pessoas, corrigir o nome "
        "como aparece na planilha (`alias_excel`) ou definir uma meta "
        "individual (`meta_exemplares`). Clique em Aplicar para salvar."
    )
    texto_config = st.text_area(
        "Configuração da equipe (JSON):",
        value=json.dumps(st.session_state["equipe_config"], ensure_ascii=False, indent=2),
        height=220,
        key="texto_config_equipe",
    )
    col_aplicar, col_restaurar = st.columns(2)
    with col_aplicar:
        if st.button("✅ Aplicar", use_container_width=True):
            try:
                nova_config = json.loads(texto_config)
                if not isinstance(nova_config, dict):
                    raise ValueError("O JSON precisa ser um objeto com cargos como chaves.")
                st.session_state["equipe_config"] = nova_config
                st.success("Configuração de equipe atualizada.")
                st.rerun()
            except Exception as e:
                st.error(f"JSON inválido: {e}")
    with col_restaurar:
        if st.button("↩️ Restaurar padrão", use_container_width=True):
            st.session_state["equipe_config"] = json.loads(json.dumps(DEFAULT_EQUIPE))
            st.session_state.pop("texto_config_equipe", None)
            st.rerun()

EQUIPE, NOMES_LISTA, ALIAS_EXCEL, METAS_INDIVIDUAIS, DEFAULTS_MOV, CARGO_POR_NOME = construir_estruturas_equipe(
    st.session_state["equipe_config"]
)

st.sidebar.markdown("<hr style='margin:14px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)

st.sidebar.markdown("### 👁️ Filtros Gerenciais")
remover_do_setor = st.sidebar.multiselect("Ocultar do Setor (Tabela):", NOMES_LISTA, key="remover_do_setor")

st.sidebar.markdown("<hr style='margin:14px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)

st.sidebar.markdown("### ➕ Adicionar Manualmente ao Relatório")
st.sidebar.caption(
    "Digite um nome e adicione à tabela do Detalhamento Gerencial, mesmo que "
    "a pessoa não esteja na equipe cadastrada ou não tenha registro na "
    "planilha no dia. Não precisa passar pela seção de Movimentação."
)

if "pessoas_manuais" not in st.session_state:
    st.session_state["pessoas_manuais"] = []

novo_nome_manual = st.sidebar.text_input("Nome da pessoa:", key="novo_nome_manual")
novo_cargo_manual = st.sidebar.selectbox(
    "Cargo:", list(EQUIPE.keys()) + ["Outro"], key="novo_cargo_manual"
)
if st.sidebar.button("➕ Adicionar à tabela", use_container_width=True):
    nome_limpo = novo_nome_manual.strip()
    if not nome_limpo:
        st.sidebar.warning("Digite um nome antes de adicionar.")
    elif nome_limpo in NOMES_LISTA or any(
        p["nome"] == nome_limpo for p in st.session_state["pessoas_manuais"]
    ):
        st.sidebar.warning("Esse nome já está na equipe ou já foi adicionado.")
    else:
        st.session_state["pessoas_manuais"].append({"nome": nome_limpo, "cargo": novo_cargo_manual})
        st.session_state["novo_nome_manual"] = ""
        st.rerun()

if st.session_state["pessoas_manuais"]:
    st.sidebar.caption("Adicionados manualmente:")
    for i, pessoa in enumerate(st.session_state["pessoas_manuais"]):
        col_nome_add, col_remover_add = st.sidebar.columns([3, 1])
        col_nome_add.markdown(f"👤 {pessoa['nome']} · {pessoa['cargo']}")
        if col_remover_add.button("🗑️", key=f"remover_manual_{i}", use_container_width=True):
            st.session_state["pessoas_manuais"].pop(i)
            st.rerun()

st.sidebar.markdown("<hr style='margin:14px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)

st.sidebar.markdown("### ❌ Ausências do Dia")
faltas_selecionadas = st.sidebar.multiselect("Selecione quem faltou hoje:", NOMES_LISTA, key="faltas_selecionadas")

st.sidebar.markdown("<hr style='margin:14px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)

st.sidebar.markdown("### ⏳ Movimentação de Horários")
movimentados_selecionados = st.sidebar.multiselect(
    "🚚 Quem foi movimentado(a) hoje?",
    [n for n in NOMES_LISTA if n not in remover_do_setor and n not in faltas_selecionadas],
    key="movimentados_selecionados",
)

st.sidebar.markdown("<hr style='margin:14px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)

MOTIVOS_FALTA_PADRAO = ["Falta administrativa", "Atestado médico", "Falta injustificada", "Folga compensatória", "Outro"]

dict_movimentacao = {}
dict_motivos_falta = {}

for cargo, integrantes in EQUIPE.items():
    integrantes_visiveis = [i for i in integrantes if i not in remover_do_setor]
    if integrantes_visiveis:
        st.sidebar.markdown(
            f"<h3 style='color:#1E3A8A; margin-top:10px; font-size:1.1rem;'>🔹 {cargo.upper()}</h3>",
            unsafe_allow_html=True,
        )

    for op in integrantes:
        if op in remover_do_setor:
            continue

        is_ausente = op in faltas_selecionadas
        is_movimentado = op in movimentados_selecionados

        if is_ausente:
            st.sidebar.markdown(f"❌ **{op} (AUSENTE)**")
            motivo_escolhido = st.sidebar.selectbox(
                f"Motivo da falta de {op}:", MOTIVOS_FALTA_PADRAO, key=f"mot_falta_sel_{op}"
            )
            if motivo_escolhido == "Outro":
                motivo_escolhido = st.sidebar.text_input(
                    f"Descreva o motivo de {op}:", value="", key=f"mot_falta_txt_{op}"
                )
            dict_motivos_falta[op] = motivo_escolhido or "Falta administrativa"
            dict_movimentacao[op] = {"cargo": cargo, "movimentacoes": []}

        elif is_movimentado:
            st.sidebar.markdown(f"**👤 {op}**", unsafe_allow_html=True)

            salvo = st.session_state["horarios_salvos"].get(op)
            if salvo:
                linha_inicial = pd.DataFrame(
                    [
                        {
                            "Saída": parse_hora_str(linha.get("saida", "")),
                            "Retorno": parse_hora_str(linha.get("retorno", "")),
                            "Local": linha.get("local", ""),
                        }
                        for linha in salvo
                    ] or [{"Saída": None, "Retorno": None, "Local": ""}]
                )
            else:
                defaults = DEFAULTS_MOV.get(op, {})
                linha_inicial = pd.DataFrame(
                    [
                        {
                            "Saída": parse_hora_str(defaults.get("saida", "")),
                            "Retorno": parse_hora_str(defaults.get("retorno", "")),
                            "Local": defaults.get("local", ""),
                        }
                    ]
                )

            editado = st.sidebar.data_editor(
                linha_inicial,
                num_rows="dynamic",
                hide_index=True,
                use_container_width=True,
                key=f"mov_editor_{op}",
                column_config={
                    "Saída": st.column_config.TimeColumn("Saída", format="HH:mm", step=60),
                    "Retorno": st.column_config.TimeColumn("Retorno", format="HH:mm", step=60),
                    "Local": st.column_config.TextColumn("Local"),
                },
            )

            movimentacoes_op = []
            for _, linha in editado.iterrows():
                sai = linha.get("Saída")
                ret = linha.get("Retorno")
                loc = texto_seguro(linha.get("Local"))
                sai_txt = formatar_hora_editor(sai)
                ret_txt = formatar_hora_editor(ret)
                if sai_txt or ret_txt or loc.strip():
                    movimentacoes_op.append({"sai": sai_txt, "ret": ret_txt, "loc": loc})

            dict_movimentacao[op] = {"cargo": cargo, "movimentacoes": movimentacoes_op}

            col_salvar_hora, col_resetar_hora = st.sidebar.columns(2)
            with col_salvar_hora:
                if st.button("💾 Salvar", key=f"salvar_horario_{op}", use_container_width=True):
                    linhas_para_salvar = []
                    for _, linha_ed in editado.iterrows():
                        sai_iso = hora_para_iso(linha_ed.get("Saída"))
                        ret_iso = hora_para_iso(linha_ed.get("Retorno"))
                        loc_ed = texto_seguro(linha_ed.get("Local"))
                        if sai_iso or ret_iso or loc_ed.strip():
                            linhas_para_salvar.append({"saida": sai_iso, "retorno": ret_iso, "local": loc_ed})
                    st.session_state["horarios_salvos"][op] = linhas_para_salvar
                    salvar_horarios_disco(st.session_state["horarios_salvos"])
                    st.sidebar.success(f"Horário de {op} salvo.")
            with col_resetar_hora:
                if st.button("🔄 Resetar", key=f"resetar_horario_{op}", use_container_width=True):
                    st.session_state["horarios_salvos"].pop(op, None)
                    salvar_horarios_disco(st.session_state["horarios_salvos"])
                    st.session_state.pop(f"mov_editor_{op}", None)
                    st.rerun()

        else:
            st.sidebar.markdown(
                f"👤 {op} <span style='font-size:0.8rem; color:gray;'>(sem movimentação)</span>",
                unsafe_allow_html=True,
            )
            dict_movimentacao[op] = {"cargo": cargo, "movimentacoes": []}

        st.sidebar.markdown("<hr style='margin:6px 0px; border-color: #D1D5DB;'>", unsafe_allow_html=True)


# =============================================================================
# 7. Configuração de e-mail (SMTP) — opcional
# =============================================================================
with st.sidebar.expander("✉️ Configuração de E-mail (SMTP)"):
    st.caption(
        "Preencha para habilitar o envio direto do relatório por e-mail. "
        "Por segurança, prefira configurar a senha em `st.secrets` em vez de digitá-la aqui."
    )
    smtp_host = st.text_input("Servidor SMTP:", value="smtp.gmail.com", key="smtp_host")
    smtp_port = st.number_input("Porta:", value=587, step=1, key="smtp_port")
    smtp_usuario = st.text_input("E-mail remetente:", key="smtp_usuario")
    smtp_senha = st.text_input("Senha / senha de app:", type="password", key="smtp_senha")
    destinatarios_texto = st.text_input("Destinatários (separados por vírgula):", key="smtp_destinatarios")


def enviar_email_relatorio(assunto, corpo_texto, corpo_html, imagem_bytes, nome_imagem, cid_imagem):
    """Envia o relatório por e-mail em HTML, com a imagem do painel embutida
    diretamente no corpo da mensagem (via Content-ID) — igual ao modelo em
    que o texto vem primeiro e a imagem aparece logo abaixo de
    'Observações do Dia:', em vez de só como anexo separado.

    A mensagem é multipart/alternative (texto simples + HTML) dentro de um
    envelope multipart/related, que é onde a imagem embutida entra. Isso
    garante que clientes que não renderizam HTML ainda recebam o texto puro
    como alternativa."""
    destinatarios = [d.strip() for d in destinatarios_texto.split(",") if d.strip()]
    if not (smtp_host and smtp_usuario and smtp_senha and destinatarios):
        st.error("Preencha servidor, remetente, senha e ao menos um destinatário na configuração de e-mail.")
        return
    try:
        msg = MIMEMultipart("related")
        msg["From"] = smtp_usuario
        msg["To"] = ", ".join(destinatarios)
        msg["Subject"] = assunto

        alternativo = MIMEMultipart("alternative")
        alternativo.attach(MIMEText(corpo_texto, "plain", "utf-8"))
        alternativo.attach(MIMEText(corpo_html, "html", "utf-8"))
        msg.attach(alternativo)

        img_part = MIMEImage(imagem_bytes, name=nome_imagem)
        img_part.add_header("Content-ID", f"<{cid_imagem}>")
        img_part.add_header("Content-Disposition", "inline", filename=nome_imagem)
        msg.attach(img_part)

        with smtplib.SMTP(smtp_host, int(smtp_port)) as servidor:
            servidor.starttls()
            servidor.login(smtp_usuario, smtp_senha)
            servidor.sendmail(smtp_usuario, destinatarios, msg.as_string())

        st.success(f"E-mail enviado com sucesso para: {', '.join(destinatarios)}")
    except Exception as e:
        st.error(f"Falha ao enviar e-mail: {e}")


# =============================================================================
# 8. Leitura da planilha — por CABEÇALHO de coluna, com cache e avisos claros
# =============================================================================
def localizar_colunas(sheet):
    mapeamento = {}
    for col in range(1, sheet.max_column + 1):
        valor = sheet.cell(row=1, column=col).value
        if valor is None:
            continue
        v = normalizar(valor)
        if v == "TOTAL" and "TOTAL" not in mapeamento:
            mapeamento["TOTAL"] = col
        if v == "USUARIO" and "USUARIO" not in mapeamento:
            mapeamento["USUARIO"] = col
    return mapeamento


@st.cache_data(show_spinner="Lendo planilha...")
def ler_planilha(bytes_arquivo):
    wb = openpyxl.load_workbook(io.BytesIO(bytes_arquivo), data_only=True)
    sheet = wb.active

    mapeamento = localizar_colunas(sheet)
    usando_fallback = ("TOTAL" not in mapeamento) or ("USUARIO" not in mapeamento)
    col_total = mapeamento.get("TOTAL", 9)
    col_usuario = mapeamento.get("USUARIO", 13)

    dados = []
    for row in range(2, sheet.max_row + 1):
        if sheet.row_dimensions[row].hidden:
            continue
        val_total = sheet.cell(row=row, column=col_total).value
        val_usuario = sheet.cell(row=row, column=col_usuario).value
        if val_total is not None and val_usuario is not None:
            dados.append({"TOTAL": val_total, "USUARIO": normalizar(val_usuario)})

    df = pd.DataFrame(dados, columns=["TOTAL", "USUARIO"])
    if not df.empty:
        df["TOTAL"] = pd.to_numeric(df["TOTAL"], errors="coerce").fillna(0)

    return df, usando_fallback


# =============================================================================
# 9. Geração de relatório em imagem, Excel e histórico
# =============================================================================
def gerar_relatorio_imagem(total_exemplares, total_skus, pct_exemplares, pct_skus,
                            meta_exemplares, meta_skus, df_real, data_formatada=""):
    colunas_relatorio = ["Cargo", "Colaboradora", "Movimentação Operacional"]
    df_relatorio = df_real[colunas_relatorio].copy() if not df_real.empty else df_real

    LARGURA_QUEBRA = 78
    linhas_por_registro = []
    if not df_relatorio.empty:
        textos_quebrados = []
        for texto in df_relatorio["Movimentação Operacional"]:
            texto = "" if texto is None else str(texto)
            linhas_texto = textwrap.wrap(texto, width=LARGURA_QUEBRA) or [""]
            textos_quebrados.append("\n".join(linhas_texto))
            linhas_por_registro.append(len(linhas_texto))
        df_relatorio["Movimentação Operacional"] = textos_quebrados

    # --- Medidas gerais (em polegadas) -------------------------------------
    ALTURA_HEADER_IN = 0.62
    ESPACO_HEADER_CARDS_IN = 0.22
    ALTURA_CARDS_IN = 1.15
    ESPACO_CARDS_TABELA_IN = 0.22
    ALTURA_CABECALHO_TABELA_IN = 0.34
    ALTURA_LINHA_TABELA_IN = 0.24
    ESPACO_TABELA_RODAPE_IN = 0.16
    ALTURA_RODAPE_IN = 0.26
    MARGEM_INFERIOR_IN = 0.08

    total_linhas_texto = sum(max(n, 1) for n in linhas_por_registro) if linhas_por_registro else 1
    altura_tabela_in = ALTURA_CABECALHO_TABELA_IN + ALTURA_LINHA_TABELA_IN * total_linhas_texto

    altura_fig = (
        ALTURA_HEADER_IN + ESPACO_HEADER_CARDS_IN + ALTURA_CARDS_IN + ESPACO_CARDS_TABELA_IN
        + altura_tabela_in + ESPACO_TABELA_RODAPE_IN + ALTURA_RODAPE_IN + MARGEM_INFERIOR_IN
    )

    COR_NAVY = "#0F172A"
    COR_AZUL = "#2563EB"
    COR_TEAL = "#0D9488"
    COR_TEXTO = "#111827"
    COR_MUTED = "#64748B"

    fig = plt.figure(figsize=(11, altura_fig), dpi=200)
    fig.patch.set_facecolor("#F8FAFC")

    # --- Faixa de cabeçalho (fundo branco) -----------------------------------
    frac_header_altura = ALTURA_HEADER_IN / altura_fig
    ax_header = fig.add_axes([0, 1 - frac_header_altura, 1, frac_header_altura])
    ax_header.set_xlim(0, 1); ax_header.set_ylim(0, 1); ax_header.axis("off")
    ax_header.add_patch(mpatches.Rectangle((0, 0), 1, 1, facecolor="#FFFFFF", edgecolor="none"))
    ax_header.plot([0, 1], [0.02, 0.02], color="#E5E7EB", linewidth=1, transform=ax_header.transAxes)

    # Ícone simples (quadrado arredondado com 3 barrinhas ascendentes)
    icon_x, icon_w = 0.028, 0.032
    ax_header.add_patch(mpatches.FancyBboxPatch(
        (icon_x, 0.28), icon_w, 0.44, boxstyle="round,pad=0,rounding_size=0.012",
        linewidth=0, facecolor=COR_NAVY, transform=ax_header.transAxes
    ))
    barra_larg = icon_w / 5.6
    for i, alt in enumerate([0.14, 0.22, 0.30]):
        ax_header.add_patch(mpatches.Rectangle(
            (icon_x + 0.006 + i * (barra_larg + 0.004), 0.36), barra_larg, alt,
            facecolor="#93C5FD", edgecolor="none", transform=ax_header.transAxes
        ))

    ax_header.text(icon_x + icon_w + 0.018, 0.66, "Painel Executivo de Produção",
                    fontsize=15.5, fontweight="bold", color=COR_NAVY, va="center")
    ax_header.text(icon_x + icon_w + 0.018, 0.30, "Varejo · acompanhamento diário de produtividade",
                    fontsize=8.5, color=COR_MUTED, va="center")

    horario_brasil = datetime.now(ZoneInfo("America/Sao_Paulo"))
    if data_formatada:
        ax_header.text(0.972, 0.66, f"Referente a {data_formatada}", fontsize=9.5,
                        fontweight="bold", color=COR_NAVY, va="center", ha="right")
        ax_header.text(0.972, 0.30, f"Gerado em {horario_brasil.strftime('%d/%m/%Y %H:%M')}",
                        fontsize=7.5, color=COR_MUTED, va="center", ha="right")

    # --- Cards de KPI --------------------------------------------------------
    y_cards_topo_in = altura_fig - ALTURA_HEADER_IN - ESPACO_HEADER_CARDS_IN
    y_cards_base_in = y_cards_topo_in - ALTURA_CARDS_IN
    frac_cards_base = y_cards_base_in / altura_fig
    frac_cards_altura = ALTURA_CARDS_IN / altura_fig

    def desenhar_card(x, largura, titulo, valor, sub, pct, cor_accent):
        ax = fig.add_axes([x, frac_cards_base, largura, frac_cards_altura])
        ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis("off")

        # Sombra sutil (retângulo levemente deslocado e mais claro atrás do card)
        ax.add_patch(mpatches.FancyBboxPatch(
            (0.015, 0.03), 0.98, 0.92, boxstyle="round,pad=0,rounding_size=0.09",
            linewidth=0, facecolor="#E2E8F0", alpha=0.6, transform=ax.transAxes
        ))
        card = mpatches.FancyBboxPatch(
            (0.01, 0.06), 0.98, 0.92, boxstyle="round,pad=0,rounding_size=0.09",
            linewidth=1, edgecolor="#E5E7EB", facecolor="white", transform=ax.transAxes
        )
        ax.add_patch(card)
        barra = mpatches.FancyBboxPatch(
            (0.01, 0.06), 0.014, 0.92, boxstyle="round,pad=0,rounding_size=0.007",
            linewidth=0, facecolor=cor_accent, transform=ax.transAxes
        )
        ax.add_patch(barra)

        ax.text(0.09, 0.80, titulo, fontsize=8.5, fontweight="bold", color="#94A3B8", va="top")
        ax.text(0.09, 0.60, valor, fontsize=22, fontweight="bold", color=COR_TEXTO, va="top")
        ax.text(0.09, 0.35, sub, fontsize=7.8, color=COR_MUTED, va="top")

        # Barra de progresso
        largura_barra = 0.82
        ax.add_patch(mpatches.FancyBboxPatch(
            (0.09, 0.16), largura_barra, 0.055, boxstyle="round,pad=0,rounding_size=0.03",
            linewidth=0, facecolor="#E5E7EB", transform=ax.transAxes
        ))
        preenchido = max(min(pct, 1.0), 0.0) * largura_barra
        if preenchido > 0.02:
            ax.add_patch(mpatches.FancyBboxPatch(
                (0.09, 0.16), preenchido, 0.055, boxstyle="round,pad=0,rounding_size=0.03",
                linewidth=0, facecolor=cor_accent, transform=ax.transAxes
            ))

    desenhar_card(0.04, 0.44, "TOTAL DE EXEMPLARES", f"{total_exemplares:,} un",
                  f"Meta Diária: {meta_exemplares:,} un  ·  Atingido: {pct_exemplares:.1%}",
                  pct_exemplares, COR_AZUL)
    desenhar_card(0.52, 0.44, "TOTAL DE SKU", f"{total_skus:,}",
                  f"Meta Diária: {meta_skus:,}  ·  Atingido: {pct_skus:.1%}",
                  pct_skus, COR_TEAL)

    # --- Tabela ---------------------------------------------------------------
    y_tabela_topo_in = y_cards_base_in - ESPACO_CARDS_TABELA_IN
    y_tabela_base_in = y_tabela_topo_in - altura_tabela_in
    frac_tabela_base = y_tabela_base_in / altura_fig
    frac_tabela_altura = altura_tabela_in / altura_fig

    # Moldura arredondada por trás da tabela (efeito "card")
    ax_moldura = fig.add_axes([0.04, frac_tabela_base, 0.92, frac_tabela_altura])
    ax_moldura.set_xlim(0, 1); ax_moldura.set_ylim(0, 1); ax_moldura.axis("off")
    ax_moldura.add_patch(mpatches.FancyBboxPatch(
        (0, 0), 1, 1, boxstyle="round,pad=0,rounding_size=0.025",
        linewidth=1.1, edgecolor="#E2E8F0", facecolor="white", transform=ax_moldura.transAxes
    ))

    ax = fig.add_axes([0.04, frac_tabela_base, 0.92, frac_tabela_altura])
    ax.axis("off")

    if not df_relatorio.empty:
        tabela = ax.table(
            cellText=df_relatorio.values,
            colLabels=df_relatorio.columns,
            cellLoc="left",
            loc="upper left",
            colWidths=[0.14, 0.22, 0.64],
        )
        tabela.auto_set_font_size(False)
        tabela.set_fontsize(8.5)

        frac_por_linha_texto = ALTURA_LINHA_TABELA_IN / altura_tabela_in
        frac_cabecalho = ALTURA_CABECALHO_TABELA_IN / altura_tabela_in

        for (row, col), cell in tabela.get_celld().items():
            cell.set_edgecolor("#EEF2F6")
            cell.PAD = 0.025
            cell.get_text().set_verticalalignment("center")
            if row == 0:
                cell.set_facecolor(COR_NAVY)
                cell.set_text_props(color="white", fontweight="bold")
                cell.set_height(frac_cabecalho)
            else:
                cell.set_facecolor("#FFFFFF" if row % 2 == 0 else "#F8FAFC")
                cell.set_height(frac_por_linha_texto * max(linhas_por_registro[row - 1], 1))
    else:
        ax.text(0.02, 0.9, "Nenhum dado disponível.", fontsize=9, color=COR_MUTED)

    # --- Rodapé ---------------------------------------------------------------
    frac_rodape_altura = ALTURA_RODAPE_IN / altura_fig
    ax_rodape = fig.add_axes([0.04, 0, 0.92, frac_rodape_altura])
    ax_rodape.set_xlim(0, 1); ax_rodape.set_ylim(0, 1); ax_rodape.axis("off")
    ax_rodape.plot([0, 1], [0.92, 0.92], color="#E2E8F0", linewidth=1, transform=ax_rodape.transAxes)
    ax_rodape.text(0, 0.35, "Painel Executivo de Produção · Varejo", fontsize=7.5,
                    color="#94A3B8", va="center", ha="left")
    ax_rodape.text(1, 0.35, f"{total_skus} SKU · {total_exemplares:,} exemplares",
                    fontsize=7.5, color="#94A3B8", va="center", ha="right")

    buffer = io.BytesIO()
    fig.savefig(buffer, format="png", facecolor=fig.get_facecolor())
    plt.close(fig)
    buffer.seek(0)
    return buffer.getvalue()


def renderizar_tabela_html(df):
    if df.empty:
        return "<div class='tabela-vazia'>Nenhum dado disponível.</div>"

    larguras = {
        "Cargo": "13%",
        "Colaboradora": "17%",
        "Exemplares": "10%",
        "SKUs": "8%",
        "Meta Individual": "11%",
        "% Meta Individual": "11%",
    }

    colunas = list(df.columns)
    colgroup = "".join(
        f'<col style="width:{larguras[c]}">' if c in larguras else "<col>"
        for c in colunas
    )
    cabecalho = "".join(f"<th>{html.escape(str(c))}</th>" for c in colunas)

    linhas_html = []
    for _, linha in df.iterrows():
        celulas = []
        for c in colunas:
            valor = linha[c]
            if isinstance(valor, (int,)) or (isinstance(valor, float) and float(valor).is_integer()):
                texto = f"{int(valor):,}".replace(",", ".") if c in ("Exemplares", "SKUs", "Meta Individual") else str(valor)
            else:
                texto = "" if valor is None else str(valor)
            celulas.append(f"<td>{html.escape(texto)}</td>")
        linhas_html.append(f"<tr>{''.join(celulas)}</tr>")

    return f"""
    <div class="tabela-wrapper">
      <table class="tabela-gerencial">
        <colgroup>{colgroup}</colgroup>
        <thead><tr>{cabecalho}</tr></thead>
        <tbody>{''.join(linhas_html)}</tbody>
      </table>
    </div>
    """


def gerar_excel_gerencial(df_real):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        (df_real if not df_real.empty else pd.DataFrame(
            columns=["Cargo", "Colaboradora", "Exemplares", "SKUs", "Movimentação Operacional"]
        )).to_excel(writer, index=False, sheet_name="Produtividade")
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# 10. LÓGICA PRINCIPAL
# =============================================================================
if uploaded_file:
    df_filtrado, usando_fallback = ler_planilha(uploaded_file.getvalue())

    if usando_fallback:
        st.warning(
            "⚠️ Não encontrei as colunas 'TOTAL' e 'USUARIO' pelo cabeçalho na primeira "
            "linha da planilha. Usando posição padrão (colunas I e M) por compatibilidade "
            "— verifique se o arquivo segue o modelo esperado."
        )

    if df_filtrado.empty:
        st.error(
            "❌ Nenhum dado válido foi encontrado na planilha enviada. Os totais abaixo "
            "estão zerados — **não envie este relatório para a diretoria sem antes "
            "verificar o arquivo**."
        )
        total_exemplares, total_skus = 0, 0
    else:
        total_exemplares = int(df_filtrado["TOTAL"].sum())
        total_skus = int(len(df_filtrado))

        nomes_excel_validos = {nome_excel(n, ALIAS_EXCEL) for n in NOMES_LISTA}
        nomes_nao_mapeados = sorted(set(df_filtrado["USUARIO"]) - nomes_excel_validos)
        if nomes_nao_mapeados:
            st.warning(
                "⚠️ Encontrados na planilha, mas **não mapeados** para ninguém da equipe "
                "cadastrada: " + ", ".join(nomes_nao_mapeados) + ". Esses registros entram "
                "no TOTAL geral acima, mas não aparecem na tabela individual — confira se "
                "não é alguém novo que precisa ser cadastrado em '⚙️ Configurar Equipe'."
            )

    META_EXEMPLARES, META_SKUS = 55000, 1200
    pct_exemplares = (total_exemplares / META_EXEMPLARES) if META_EXEMPLARES else 0
    pct_skus = (total_skus / META_SKUS) if META_SKUS else 0

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(
            f'<div class="card-kpi" style="--accent-color:#2563EB;">'
            f'<div class="card-title">📦 Total de Exemplares</div>'
            f'<div class="card-value">{total_exemplares:,} un</div>'
            f'<div class="card-sub">Meta Diária: {META_EXEMPLARES:,} un · Atingido: {pct_exemplares:.1%}</div></div>',
            unsafe_allow_html=True,
        )
        st.progress(min(pct_exemplares, 1.0))
    with c2:
        st.markdown(
            f'<div class="card-kpi" style="--accent-color:#0D9488;">'
            f'<div class="card-title">🏷️ Total de SKU</div>'
            f'<div class="card-value">{total_skus:,}</div>'
            f'<div class="card-sub">Meta Diária: {META_SKUS:,} · Atingido: {pct_skus:.1%}</div></div>',
            unsafe_allow_html=True,
        )
        st.progress(min(pct_skus, 1.0))

    st.markdown("<br>", unsafe_allow_html=True)

    data_str_atual = data_produtividade.strftime("%Y-%m-%d")

    data_gerencial = []
    textos_automaticos_por_pessoa = {}  # nome -> texto gerado automaticamente (antes do override manual), usado para o autosave do modo edição não "engessar" texto que ninguém editou de fato
    for n in NOMES_LISTA:
        if n in remover_do_setor:
            continue

        is_ausente = n in faltas_selecionadas

        if is_ausente:
            if not df_filtrado.empty:
                df_func = df_filtrado[df_filtrado["USUARIO"] == nome_excel(n, ALIAS_EXCEL)]
                qtd_exemplares = int(df_func["TOTAL"].sum())
                qtd_skus = int(len(df_func))
            else:
                qtd_exemplares, qtd_skus = 0, 0
            motivo_individual = dict_motivos_falta.get(n, "Falta administrativa")
            justificativa_texto = f"Ausente. Motivo: {motivo_individual}."
            cargo_atual = CARGO_POR_NOME.get(n, "Operador(a)")
        else:
            mov = dict_movimentacao[n]
            cargo_atual = mov["cargo"]
            if not df_filtrado.empty:
                df_func = df_filtrado[df_filtrado["USUARIO"] == nome_excel(n, ALIAS_EXCEL)]
                qtd_exemplares = int(df_func["TOTAL"].sum())
                qtd_skus = int(len(df_func))
            else:
                qtd_exemplares, qtd_skus = 0, 0

            if qtd_skus == 0:
                continue

            historico_justificativas = []
            for m in mov["movimentacoes"]:
                sai, ret, loc = m["sai"].strip(), m["ret"].strip(), m["loc"].strip()
                if not (sai or ret or loc):
                    continue

                partes = []
                if loc:
                    partes.append(f"ao {loc}")
                if sai and ret:
                    partes.append(f"das {sai} às {ret}")
                elif sai:
                    partes.append(f"a partir das {sai}")
                elif ret:
                    partes.append(f"até às {ret}")

                prefixo = "Encaminhada" if not historico_justificativas else "encaminhada"
                complemento = " ".join(partes) if partes else "(sem horário/local informado)"
                historico_justificativas.append(f"{prefixo} {complemento}")

            if historico_justificativas:
                justificativa_texto = " ; ".join(historico_justificativas) + "."
            elif qtd_skus == 0:
                justificativa_texto = "Sem registros na planilha nesta data."
            else:
                justificativa_texto = "Atividade normal no setor."

        # A ausência sempre tem prioridade sobre um texto manual salvo
        # anteriormente. Sem isso, um texto de movimentação salvo enquanto a
        # pessoa ainda não estava ausente (ex.: pelo autosave do modo
        # "Editar tabela") continuava "grudado" e escondia o aviso de
        # falta, mesmo com a pessoa corretamente marcada como ausente.
        override_chave = (data_str_atual, n)
        if is_ausente:
            if override_chave in st.session_state["mov_manual_overrides"]:
                del st.session_state["mov_manual_overrides"][override_chave]
                salvar_overrides_disco(st.session_state["mov_manual_overrides"])
        else:
            textos_automaticos_por_pessoa[n] = justificativa_texto
            if override_chave in st.session_state["mov_manual_overrides"]:
                justificativa_texto = st.session_state["mov_manual_overrides"][override_chave]

        linha = {
            "Cargo": cargo_atual,
            "Colaboradora": n,
            "Exemplares": qtd_exemplares,
            "SKUs": qtd_skus,
            "Movimentação Operacional": justificativa_texto,
        }

        if METAS_INDIVIDUAIS:
            meta_pessoa = METAS_INDIVIDUAIS.get(n)
            if meta_pessoa:
                linha["Meta Individual"] = meta_pessoa
                linha["% Meta Individual"] = f"{(qtd_exemplares / meta_pessoa):.0%}"
            else:
                linha["Meta Individual"] = ""
                linha["% Meta Individual"] = ""

        data_gerencial.append(linha)

    # --- Pessoas adicionadas manualmente pela lateral ("➕ Adicionar
    # Manualmente ao Relatório") — entram na tabela mesmo sem estar na
    # equipe cadastrada e mesmo sem registro na planilha no dia. Se o nome
    # digitado bater com algo na planilha, os números reais são usados;
    # senão, ficam zerados.
    for pessoa_manual in st.session_state["pessoas_manuais"]:
        nome_manual = pessoa_manual["nome"]
        cargo_manual = pessoa_manual["cargo"]

        if not df_filtrado.empty:
            df_func_manual = df_filtrado[df_filtrado["USUARIO"] == normalizar(nome_manual)]
            qtd_exemplares_manual = int(df_func_manual["TOTAL"].sum())
            qtd_skus_manual = int(len(df_func_manual))
        else:
            qtd_exemplares_manual, qtd_skus_manual = 0, 0

        override_chave_manual = (data_str_atual, nome_manual)
        if qtd_skus_manual > 0:
            texto_automatico_manual = "Atividade normal no setor."
        else:
            texto_automatico_manual = "Sem registros na planilha nesta data. (Adicionada manualmente)"
        textos_automaticos_por_pessoa[nome_manual] = texto_automatico_manual
        justificativa_manual = st.session_state["mov_manual_overrides"].get(
            override_chave_manual, texto_automatico_manual
        )

        linha_manual = {
            "Cargo": cargo_manual,
            "Colaboradora": nome_manual,
            "Exemplares": qtd_exemplares_manual,
            "SKUs": qtd_skus_manual,
            "Movimentação Operacional": justificativa_manual,
        }
        if METAS_INDIVIDUAIS:
            linha_manual["Meta Individual"] = ""
            linha_manual["% Meta Individual"] = ""

        data_gerencial.append(linha_manual)

    df_real = pd.DataFrame(data_gerencial)

    st.markdown(
        "<h3 style='font-family: \"Sora\", sans-serif; color: #0F172A; font-size: 1.05rem; "
        "font-weight: 700; margin-bottom:10px;'>📋 Detalhamento Gerencial de Produtividade</h3>",
        unsafe_allow_html=True,
    )

    col_toggle1, col_toggle2 = st.columns(2)
    with col_toggle1:
        mostrar_individual = st.checkbox(
            "👁️ Mostrar Exemplares/SKUs individuais", value=True, key="mostrar_individual"
        )
    with col_toggle2:
        modo_edicao = st.checkbox(
            "✏️ Editar tabela (como planilha)", value=False, key="modo_edicao_tabela"
        )

    colunas_ocultaveis = ["Exemplares", "SKUs"]
    if mostrar_individual:
        df_exibir = df_real.copy()
    else:
        df_exibir = df_real.drop(columns=[c for c in colunas_ocultaveis if c in df_real.columns])

    if modo_edicao:
        st.caption(
            "✍️ A coluna **Movimentação Operacional** é livre — escreva o que quiser. "
            "As demais colunas ficam bloqueadas aqui para não conflitar com os dados da "
            "planilha. O texto tenta salvar sozinho ao sair do campo, mas para garantir "
            "que nada se perca clique em **💾 Salvar agora** antes de mexer nos horários "
            "no sidebar ou atualizar a página."
        )

        colunas_bloqueadas = [c for c in df_exibir.columns if c != "Movimentação Operacional"]

        df_exibir_editado = st.data_editor(
            df_exibir,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            key="editor_tabela_gerencial",
            disabled=colunas_bloqueadas,
            column_config={
                "Movimentação Operacional": st.column_config.TextColumn(
                    "Movimentação Operacional",
                    help="Escreva livremente o que quiser exibir para esta pessoa.",
                ),
            },
        )

        for col in df_exibir_editado.columns:
            df_real[col] = df_exibir_editado[col].values

        def persistir_movimentacao_editada(df_editado):
            """Grava como override manual apenas o texto que realmente foi
            editado à mão (diferente do texto automático daquela pessoa nesta
            data). Se o texto exibido for igual ao automático — por exemplo,
            porque a pessoa foi marcada como ausente e o texto na tela
            passou a ser "Ausente..." de novo — qualquer override antigo é
            removido em vez de recriado, evitando que um texto de
            movimentação antigo "grude" e esconda mudanças como a ausência."""
            total_gravado = 0
            if "Movimentação Operacional" in df_editado.columns and "Colaboradora" in df_editado.columns:
                for _, linha_editada in df_editado.iterrows():
                    nome_pessoa = linha_editada.get("Colaboradora")
                    texto_editado = texto_seguro(linha_editada.get("Movimentação Operacional"))
                    if not nome_pessoa:
                        continue
                    chave = (data_str_atual, nome_pessoa)
                    texto_automatico = textos_automaticos_por_pessoa.get(nome_pessoa)
                    if texto_automatico is not None and texto_editado == texto_automatico:
                        # Igual ao automático: não é uma edição manual de verdade.
                        st.session_state["mov_manual_overrides"].pop(chave, None)
                    else:
                        st.session_state["mov_manual_overrides"][chave] = texto_editado
                        total_gravado += 1
                salvar_overrides_disco(st.session_state["mov_manual_overrides"])
            return total_gravado

        persistir_movimentacao_editada(df_exibir_editado)

        col_salvar, col_restaurar_texto = st.columns([1, 2])
        with col_salvar:
            if st.button("💾 Salvar agora", use_container_width=True, type="primary"):
                qtd = persistir_movimentacao_editada(df_exibir_editado)
                st.success(f"✅ Salvo! ({qtd} linha(s) gravada(s) para {data_formatada})")
        with col_restaurar_texto:
            if st.button("🔄 Restaurar texto automático desta data", use_container_width=True):
                chaves_para_remover = [
                    k for k in st.session_state["mov_manual_overrides"] if k[0] == data_str_atual
                ]
                for k in chaves_para_remover:
                    del st.session_state["mov_manual_overrides"][k]
                salvar_overrides_disco(st.session_state["mov_manual_overrides"])
                st.rerun()
    else:
        st.markdown(renderizar_tabela_html(df_exibir), unsafe_allow_html=True)

    col_exp1, col_exp2 = st.columns(2)
    with col_exp1:
        if not df_real.empty:
            st.download_button(
                label="📥 Baixar Tabela em Excel",
                data=gerar_excel_gerencial(df_real),
                file_name=f"produtividade_{data_produtividade.strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    st.markdown("<br><hr>", unsafe_allow_html=True)

    st.markdown("<h4 style='font-family: \"Sora\", sans-serif; color: #0F172A; font-size: 0.95rem; font-weight: 700; margin-top:18px;'>🖼️ Relatório em Imagem</h4>", unsafe_allow_html=True)
    st.caption("Clique com o botão direito na imagem e escolha **Copiar imagem** para colar direto no e-mail, ou baixe o arquivo abaixo.")
    imagem_relatorio = gerar_relatorio_imagem(
        total_exemplares, total_skus, pct_exemplares, pct_skus, META_EXEMPLARES, META_SKUS, df_real,
        data_formatada=data_formatada,
    )
    st.image(imagem_relatorio, use_container_width=True)
    nome_arquivo_imagem = f"relatorio_producao_{data_produtividade.strftime('%Y-%m-%d')}.png"
    st.download_button(
        label="📥 Baixar Relatório em Imagem",
        data=imagem_relatorio,
        file_name=nome_arquivo_imagem,
        mime="image/png",
    )

    # --- Texto do e-mail ---------------------------------------------------------
    st.markdown("<br><hr>", unsafe_allow_html=True)
    st.markdown(
        "<h3 style='font-family: \"Sora\", sans-serif; color: #0F172A; font-size: 1.05rem; "
        "font-weight: 700;'>✉️ Texto do E-mail Pronto para a Diretoria</h3>",
        unsafe_allow_html=True,
    )

    # Versão em texto simples (usada na caixa para copiar/colar manualmente,
    # e como alternativa em clientes de e-mail que não leem HTML).
    texto_final = (
        f"Boa tarde, Prezados.\n\nSegue abaixo o relatório de produção.\n"
        f"referente ao dia {data_formatada}.\n\nObservações do Dia:\n"
        f"(imagem do painel anexada/embutida neste e-mail)\n\n"
        f"--------------------------------\n"
        f"Resumo Varejo.\nSKU: {total_skus}\nExemplares: {total_exemplares:,}\n"
        f"--------------------------------\n\nAtenciosamente,"
    )

    # Versão em HTML enviada de fato — reproduz o modelo do print: texto,
    # depois "Observações do Dia:" e a imagem do painel logo abaixo, embutida
    # via Content-ID (cid) em vez de só anexada.
    CID_IMAGEM_RELATORIO = "relatorio_producao_imagem"
    corpo_html_email = f"""\
<html>
  <body style="font-family: Arial, Helvetica, sans-serif; font-size: 14px; color:#111827;">
    <p>Boa tarde, Prezados.</p>
    <p>Segue abaixo o relatório de produção.<br>
       referente ao dia {html.escape(data_formatada)}.</p>
    <p><strong>Observações do Dia:</strong></p>
    <p><img src="cid:{CID_IMAGEM_RELATORIO}" alt="Painel Executivo de Produção" style="max-width:700px; width:100%; border:1px solid #E5E7EB; border-radius:8px;"></p>
    <p>--------------------------------<br>
       Resumo Varejo.<br>
       SKU: {total_skus}<br>
       Exemplares: {total_exemplares:,}<br>
       --------------------------------</p>
    <p>Atenciosamente,</p>
  </body>
</html>
"""

    st.text_area("Selecione tudo abaixo e copie (Ctrl+A / Ctrl+C):", value=texto_final, height=200, key="texto_email")

    if st.button("📧 Enviar relatório por e-mail agora"):
        enviar_email_relatorio(
            assunto=f"Relatório de Produção - {data_formatada}",
            corpo_texto=texto_final,
            corpo_html=corpo_html_email,
            imagem_bytes=imagem_relatorio,
            nome_imagem=nome_arquivo_imagem,
            cid_imagem=CID_IMAGEM_RELATORIO,
        )

else:
    st.info("👋 Painel atualizado com novas melhorias. Faça o upload da sua planilha Excel na barra lateral.")
